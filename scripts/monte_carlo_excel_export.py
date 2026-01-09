"""
Monte Carlo Excel Export - Finance Guru™
Generates comprehensive Excel workbook with charts for analysis

Author: Dr. Priya Desai (Quant Analyst)
Date: 2026-01-02
"""

import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from pathlib import Path

# Paths
BASE_DIR = Path("/Users/ossieirondi/Documents/Irondi-Household/family-office")
JSON_PATH = BASE_DIR / "docs/fin-guru/analysis/monte-carlo-v3-2026-01-02.json"
CSV_PATH = BASE_DIR / "docs/fin-guru/analysis/monte-carlo-v3-full-results-2026-01-02.csv"
OUTPUT_PATH = BASE_DIR / "docs/fin-guru/analysis/monte-carlo-v3-analysis-2026-01-02.xlsx"

# Style definitions
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
METRIC_FILL = PatternFill("solid", fgColor="D6EAF8")
GOOD_FILL = PatternFill("solid", fgColor="D5F5E3")
WARN_FILL = PatternFill("solid", fgColor="FCF3CF")
BAD_FILL = PatternFill("solid", fgColor="FADBD8")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def load_data():
    with open(JSON_PATH, 'r') as f:
        summary = json.load(f)
    scenarios = pd.read_csv(CSV_PATH)
    return summary, scenarios

def create_summary_sheet(wb, summary):
    ws = wb.active
    ws.title = "Summary Dashboard"

    # Title
    ws['A1'] = "MONTE CARLO SIMULATION v2.0"
    ws['A1'].font = Font(bold=True, size=18, color="1F4E79")
    ws.merge_cells('A1:E1')

    ws['A2'] = "5-Bucket Income + GOOGL Scale-In + SQQQ Hedge Strategy"
    ws['A2'].font = Font(italic=True, size=12, color="666666")
    ws.merge_cells('A2:E2')

    ws['A3'] = f"Simulation Date: {summary['simulation_date']} | Model Version: {summary['model_version']} | Scenarios: {summary['n_scenarios']:,}"
    ws['A3'].font = Font(size=10, color="888888")
    ws.merge_cells('A3:E3')

    # Success Metrics Section
    row = 5
    ws[f'A{row}'] = "SUCCESS METRICS"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:C{row}')

    metrics = [
        ("Probability of $100k+ Annual Income", summary['success_metrics']['probability_100k_income'], "pct"),
        ("Probability of $75k+ Annual Income", summary['success_metrics']['probability_75k_income'], "pct"),
        ("Probability of $50k+ Annual Income", summary['success_metrics']['probability_50k_income'], "pct"),
        ("Probability of Margin-Free by Month 28", summary['success_metrics']['probability_margin_free'], "pct"),
        ("Margin Call Rate", summary['success_metrics']['margin_call_rate'], "pct"),
        ("Business Backstop Usage Rate", summary['success_metrics']['backstop_usage_rate'], "pct"),
        ("Average Backstop Amount", summary['success_metrics']['avg_backstop_amount'], "currency"),
    ]

    row += 1
    for name, value, fmt in metrics:
        ws[f'A{row}'] = name
        ws[f'A{row}'].fill = METRIC_FILL
        if fmt == "pct":
            ws[f'C{row}'] = value
            ws[f'C{row}'].number_format = '0.0%'
            if value >= 0.7:
                ws[f'C{row}'].fill = GOOD_FILL
            elif value >= 0.3:
                ws[f'C{row}'].fill = WARN_FILL
            else:
                ws[f'C{row}'].fill = BAD_FILL
        else:
            ws[f'C{row}'] = value
            ws[f'C{row}'].number_format = '$#,##0'
        row += 1

    # Portfolio Value Section
    row += 1
    ws[f'A{row}'] = "PORTFOLIO VALUE AT MONTH 28"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:C{row}')

    pv = summary['portfolio_value']
    portfolio_metrics = [
        ("Median", pv['median']),
        ("Mean", pv['mean']),
        ("5th Percentile (Bear)", pv['p5']),
        ("25th Percentile", pv['p25']),
        ("75th Percentile", pv['p75']),
        ("95th Percentile (Bull)", pv['p95']),
        ("Minimum", pv['min']),
        ("Maximum", pv['max']),
    ]

    row += 1
    for name, value in portfolio_metrics:
        ws[f'A{row}'] = name
        ws[f'A{row}'].fill = METRIC_FILL
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = '$#,##0'
        row += 1

    # Portfolio Composition Section (v3.0 - All 4 Layers)
    row += 1
    ws[f'A{row}'] = "PORTFOLIO COMPOSITION AT MONTH 28"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:C{row}')

    composition = [
        ("Layer 1 (Growth - PLTR, TSLA, etc.)", summary.get('layer1_growth', {}).get('median', 0)),
        ("Layer 2 (Income Portfolio)", summary['income_portfolio']['median']),
        ("Layer 3 (Hedge - SQQQ)", summary['hedge_position']['median']),
        ("GOOGL Position", summary['googl_position']['median']),
    ]

    row += 1
    for name, value in composition:
        ws[f'A{row}'] = name
        ws[f'A{row}'].fill = METRIC_FILL
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = '$#,##0'
        row += 1

    # Margin Ratio Section (NEW in v3.0)
    if 'margin_ratio' in summary:
        row += 1
        ws[f'A{row}'] = "MARGIN RATIO (Portfolio:Margin)"
        ws[f'A{row}'].font = HEADER_FONT
        ws[f'A{row}'].fill = HEADER_FILL
        ws.merge_cells(f'A{row}:C{row}')

        margin_ratio = summary['margin_ratio']
        ratio_metrics = [
            ("Median Ratio", margin_ratio['median'], "ratio"),
            ("5th Percentile (Worst)", margin_ratio['p5'], "ratio"),
            ("95th Percentile (Best)", margin_ratio['p95'], "ratio"),
            ("Minimum Required", 3.0, "ratio"),
        ]

        row += 1
        for name, value, fmt in ratio_metrics:
            ws[f'A{row}'] = name
            ws[f'A{row}'].fill = METRIC_FILL
            if fmt == "ratio" and value != float('inf'):
                ws[f'C{row}'] = f"{value:.2f}:1"
                if name == "Minimum Required":
                    ws[f'C{row}'].fill = WARN_FILL
                elif value >= 6:
                    ws[f'C{row}'].fill = GOOD_FILL
                elif value >= 3:
                    ws[f'C{row}'].fill = WARN_FILL
                else:
                    ws[f'C{row}'].fill = BAD_FILL
            row += 1

    # Dividend Income Section
    row += 1
    ws[f'A{row}'] = "ANNUAL DIVIDEND INCOME AT MONTH 28"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:C{row}')

    div = summary['dividend_income']
    div_metrics = [
        ("Median", div['median']),
        ("Mean", div['mean']),
        ("5th Percentile", div['p5']),
        ("25th Percentile", div['p25']),
        ("75th Percentile", div['p75']),
        ("95th Percentile", div['p95']),
    ]

    row += 1
    for name, value in div_metrics:
        ws[f'A{row}'] = name
        ws[f'A{row}'].fill = METRIC_FILL
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = '$#,##0'
        row += 1

    # Margin Balance Section
    row += 1
    ws[f'A{row}'] = "MARGIN BALANCE AT MONTH 28"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:C{row}')

    margin = summary['margin_balance']
    margin_metrics = [
        ("Median", margin['median']),
        ("Mean", margin['mean']),
        ("Maximum", margin['max']),
    ]

    row += 1
    for name, value in margin_metrics:
        ws[f'A{row}'] = name
        ws[f'A{row}'].fill = METRIC_FILL
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = '$#,##0'
        ws[f'C{row}'].fill = BAD_FILL
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 20

def create_scenario_data_sheet(wb, scenarios):
    ws = wb.create_sheet("Scenario Data")

    # Add headers with formatting
    headers = list(scenarios.columns)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    # Add data
    for r_idx, row in enumerate(scenarios.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if c_idx in [1, 2, 3, 4, 5, 6, 13]:
                cell.number_format = '$#,##0'
            elif c_idx == 7:
                cell.number_format = '0.0%'

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}{len(scenarios) + 1}"

def create_distribution_sheet(wb, scenarios):
    ws = wb.create_sheet("Distribution Analysis")

    # Title
    ws['A1'] = "DIVIDEND INCOME DISTRIBUTION ANALYSIS"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells('A1:D1')

    # Create histogram buckets for dividend income
    div_income = scenarios['final_annual_dividend']

    buckets = [
        (0, 25000, "$0 - $25k"),
        (25000, 40000, "$25k - $40k"),
        (40000, 50000, "$40k - $50k"),
        (50000, 60000, "$50k - $60k"),
        (60000, 70000, "$60k - $70k"),
        (70000, 80000, "$70k - $80k"),
        (80000, 100000, "$80k - $100k"),
        (100000, float('inf'), "$100k+"),
    ]

    ws['A3'] = "Income Range"
    ws['B3'] = "Scenarios"
    ws['C3'] = "Percentage"
    ws['D3'] = "Cumulative %"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}3'].font = HEADER_FONT
        ws[f'{col}3'].fill = HEADER_FILL

    cumulative = 0
    row = 4
    for low, high, label in buckets:
        count = ((div_income >= low) & (div_income < high)).sum()
        pct = count / len(div_income)
        cumulative += pct

        ws[f'A{row}'] = label
        ws[f'B{row}'] = count
        ws[f'C{row}'] = pct
        ws[f'C{row}'].number_format = '0.0%'
        ws[f'D{row}'] = cumulative
        ws[f'D{row}'].number_format = '0.0%'

        if low >= 50000:
            ws[f'A{row}'].fill = GOOD_FILL
        elif low >= 40000:
            ws[f'A{row}'].fill = WARN_FILL
        else:
            ws[f'A{row}'].fill = BAD_FILL
        row += 1

    # Portfolio Value Distribution
    row += 2
    ws[f'A{row}'] = "PORTFOLIO VALUE DISTRIBUTION"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells(f'A{row}:D{row}')

    pv = scenarios['final_portfolio_value']
    pv_buckets = [
        (0, 300000, "$0 - $300k"),
        (300000, 350000, "$300k - $350k"),
        (350000, 400000, "$350k - $400k"),
        (400000, 450000, "$400k - $450k"),
        (450000, 500000, "$450k - $500k"),
        (500000, float('inf'), "$500k+"),
    ]

    row += 2
    ws[f'A{row}'] = "Portfolio Range"
    ws[f'B{row}'] = "Scenarios"
    ws[f'C{row}'] = "Percentage"
    ws[f'D{row}'] = "Cumulative %"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = HEADER_FONT
        ws[f'{col}{row}'].fill = HEADER_FILL

    cumulative = 0
    row += 1
    pv_data_start = row
    for low, high, label in pv_buckets:
        count = ((pv >= low) & (pv < high)).sum()
        pct = count / len(pv)
        cumulative += pct

        ws[f'A{row}'] = label
        ws[f'B{row}'] = count
        ws[f'C{row}'] = pct
        ws[f'C{row}'].number_format = '0.0%'
        ws[f'D{row}'] = cumulative
        ws[f'D{row}'].number_format = '0.0%'
        row += 1

    # Key Statistics Table
    row += 2
    ws[f'A{row}'] = "KEY STATISTICS"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells(f'A{row}:D{row}')

    row += 2
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Dividend Income"
    ws[f'C{row}'] = "Portfolio Value"
    ws[f'D{row}'] = "Margin Balance"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = HEADER_FONT
        ws[f'{col}{row}'].fill = HEADER_FILL

    stats = [
        ("Mean", div_income.mean(), pv.mean(), scenarios['final_margin_balance'].mean()),
        ("Median", div_income.median(), pv.median(), scenarios['final_margin_balance'].median()),
        ("Std Dev", div_income.std(), pv.std(), scenarios['final_margin_balance'].std()),
        ("Min", div_income.min(), pv.min(), scenarios['final_margin_balance'].min()),
        ("Max", div_income.max(), pv.max(), scenarios['final_margin_balance'].max()),
    ]

    row += 1
    for stat_name, div_val, pv_val, margin_val in stats:
        ws[f'A{row}'] = stat_name
        ws[f'B{row}'] = div_val
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'C{row}'] = pv_val
        ws[f'C{row}'].number_format = '$#,##0'
        ws[f'D{row}'] = margin_val
        ws[f'D{row}'].number_format = '$#,##0'
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    return 4  # Return dividend chart data start row

def create_charts_sheet(wb, scenarios):
    ws = wb.create_sheet("Charts")

    # Title
    ws['A1'] = "MONTE CARLO VISUALIZATION"
    ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws.merge_cells('A1:H1')

    # Prepare histogram data for dividend income
    div_income = scenarios['final_annual_dividend']

    buckets = [
        ("$0-25k", 0, 25000),
        ("$25-40k", 25000, 40000),
        ("$40-50k", 40000, 50000),
        ("$50-60k", 50000, 60000),
        ("$60-70k", 60000, 70000),
        ("$70-80k", 70000, 80000),
        ("$80-100k", 80000, 100000),
        ("$100k+", 100000, float('inf')),
    ]

    # Write chart data
    ws['A3'] = "Dividend Income Distribution"
    ws['A3'].font = Font(bold=True, size=12)

    ws['A4'] = "Range"
    ws['B4'] = "Count"
    ws['C4'] = "Percentage"
    for col in ['A', 'B', 'C']:
        ws[f'{col}4'].font = HEADER_FONT
        ws[f'{col}4'].fill = HEADER_FILL

    row = 5
    for label, low, high in buckets:
        count = ((div_income >= low) & (div_income < high)).sum()
        ws[f'A{row}'] = label
        ws[f'B{row}'] = count
        ws[f'C{row}'] = count / len(div_income)
        ws[f'C{row}'].number_format = '0.0%'
        row += 1

    # Create bar chart for dividend distribution
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Dividend Income Distribution (10,000 Scenarios)"
    chart1.y_axis.title = "Number of Scenarios"
    chart1.x_axis.title = "Annual Dividend Income"

    data = Reference(ws, min_col=2, min_row=4, max_row=12, max_col=2)
    cats = Reference(ws, min_col=1, min_row=5, max_row=12)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    chart1.width = 18
    chart1.height = 10

    ws.add_chart(chart1, "E3")

    # Portfolio Value Distribution Chart
    pv = scenarios['final_portfolio_value']
    pv_buckets = [
        ("$200-300k", 200000, 300000),
        ("$300-350k", 300000, 350000),
        ("$350-400k", 350000, 400000),
        ("$400-450k", 400000, 450000),
        ("$450-500k", 450000, 500000),
        ("$500k+", 500000, float('inf')),
    ]

    ws['A15'] = "Portfolio Value Distribution"
    ws['A15'].font = Font(bold=True, size=12)

    ws['A16'] = "Range"
    ws['B16'] = "Count"
    ws['C16'] = "Percentage"
    for col in ['A', 'B', 'C']:
        ws[f'{col}16'].font = HEADER_FONT
        ws[f'{col}16'].fill = HEADER_FILL

    row = 17
    for label, low, high in pv_buckets:
        count = ((pv >= low) & (pv < high)).sum()
        ws[f'A{row}'] = label
        ws[f'B{row}'] = count
        ws[f'C{row}'] = count / len(pv)
        ws[f'C{row}'].number_format = '0.0%'
        row += 1

    # Create bar chart for portfolio value
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 11
    chart2.title = "Portfolio Value Distribution (10,000 Scenarios)"
    chart2.y_axis.title = "Number of Scenarios"
    chart2.x_axis.title = "Portfolio Value at Month 28"

    data2 = Reference(ws, min_col=2, min_row=16, max_row=22, max_col=2)
    cats2 = Reference(ws, min_col=1, min_row=17, max_row=22)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.shape = 4
    chart2.width = 18
    chart2.height = 10

    ws.add_chart(chart2, "E15")

    # Percentile comparison chart
    ws['A27'] = "Percentile Comparison"
    ws['A27'].font = Font(bold=True, size=12)

    ws['A28'] = "Percentile"
    ws['B28'] = "Dividend Income"
    ws['C28'] = "Portfolio Value"
    for col in ['A', 'B', 'C']:
        ws[f'{col}28'].font = HEADER_FONT
        ws[f'{col}28'].fill = HEADER_FILL

    percentiles = [5, 10, 25, 50, 75, 90, 95]
    row = 29
    for p in percentiles:
        ws[f'A{row}'] = f"{p}th"
        ws[f'B{row}'] = div_income.quantile(p/100)
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'C{row}'] = pv.quantile(p/100)
        ws[f'C{row}'].number_format = '$#,##0'
        row += 1

    # Create line chart for percentiles
    chart3 = LineChart()
    chart3.style = 12
    chart3.title = "Outcome Percentiles"
    chart3.y_axis.title = "Value ($)"
    chart3.x_axis.title = "Percentile"

    data3 = Reference(ws, min_col=2, min_row=28, max_row=35, max_col=3)
    cats3 = Reference(ws, min_col=1, min_row=29, max_row=35)
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.width = 18
    chart3.height = 10

    ws.add_chart(chart3, "E27")

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18

def create_assumptions_sheet(wb):
    ws = wb.create_sheet("Model Assumptions")

    ws['A1'] = "MONTE CARLO MODEL v3.0 ASSUMPTIONS (Full Portfolio)"
    ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws.merge_cells('A1:C1')

    assumptions = [
        ("STARTING POSITIONS (Jan 2, 2026)", "", ""),
        ("Layer 1 (Growth)", "$170,073", "PLTR, TSLA, VOO, SPMO, etc."),
        ("Layer 2 (Income)", "$61,725", "JEPI, JEPQ, CLM, CRF, etc."),
        ("Layer 3 (Hedge)", "$13,199", "SQQQ position"),
        ("GOOGL", "$1,876", "6.004 shares"),
        ("Starting Margin Debt", "$3,222", ""),
        ("TOTAL PORTFOLIO", "$246,873", ""),
        ("", "", ""),
        ("MONTHLY DEPLOYMENT", "", ""),
        ("Monthly W2 Income", "$13,317", ""),
        ("Layer 2 Deployment", "$11,517/month", "94% to income"),
        ("GOOGL Scale-In", "$1,000/month", "Diverted from MSTY"),
        ("Layer 3 Hedge (SQQQ)", "$800/month", "6% allocation"),
        ("Layer 1 Deployment", "$0/month", "Keep 100% - no new adds"),
        ("", "", ""),
        ("INCOME ALLOCATION (5 BUCKETS)", "", ""),
        ("JPMorgan Income (JEPI/JEPQ)", "27%", "9% yield"),
        ("CEF Stable (CLM/CRF/ECAT)", "20%", "21% yield"),
        ("Covered Call ETFs (QQQI/SPYI/QQQY)", "35%", "16% yield"),
        ("YieldMax (YMAX/AMZY)", "10%", "70% yield - MSTY paused"),
        ("DRIP v2 CEFs (BDJ/ETY/ETV/BST/UTG)", "8%", "9% yield"),
        ("Blended Target Yield", "19.9%", "Weighted average"),
        ("", "", ""),
        ("MARGIN PARAMETERS", "", ""),
        ("Starting Margin Balance", "$3,222", "Jan 2, 2026"),
        ("Margin Rate", "11.825%", "Fidelity rate"),
        ("Months 1-6 Draw", "$4,500/month", "Fixed expenses"),
        ("Months 7-12 Draw", "$6,213/month", "Add mortgage"),
        ("Months 13-28 Draw", "$8,000/month", "Partial variables"),
        ("Min Portfolio:Margin Ratio", "3:1", "Trigger backstop"),
        ("Business Income Backstop", "$22,000/month", "Insurance"),
        ("", "", ""),
        ("MARKET ASSUMPTIONS", "", ""),
        ("Bull Market Probability", "35%", "+15% annual"),
        ("Normal Market Probability", "40%", "+8% annual"),
        ("Bear Market Probability", "20%", "-15% annual"),
        ("Crisis Probability", "5%", "-40% annual"),
        ("Yield Degradation", "15%", "Over 28 months"),
        ("", "", ""),
        ("LAYER CHARACTERISTICS", "", ""),
        ("Layer 1 Beta", "1.2", "High market sensitivity"),
        ("Layer 1 Volatility", "22%", "Annual"),
        ("Layer 1 Expected Return", "18%", "Annual (growth stocks)"),
        ("Layer 2 Volatility", "6.9% - 45%", "Varies by bucket"),
        ("Layer 3 (SQQQ)", "-3x market", "With volatility decay"),
        ("GOOGL Volatility", "30%", "Single stock risk"),
        ("", "", ""),
        ("MARGIN CALCULATION", "", ""),
        ("Margin Ratio Formula", "TOTAL Portfolio / Margin", "Includes ALL layers"),
        ("Minimum Ratio Required", "3.0:1", "Margin call trigger"),
        ("Simulation Period", "28 months", ""),
        ("Scenarios Run", "10,000", ""),
    ]

    ws['A3'] = "Parameter"
    ws['B3'] = "Value"
    ws['C3'] = "Notes"
    for col in ['A', 'B', 'C']:
        ws[f'{col}3'].font = HEADER_FONT
        ws[f'{col}3'].fill = HEADER_FILL

    row = 4
    for param, value, notes in assumptions:
        if param == "":
            row += 1
            continue
        if value == "" and notes == "":
            ws[f'A{row}'] = param
            ws[f'A{row}'].font = Font(bold=True, size=11, color="1F4E79")
            ws.merge_cells(f'A{row}:C{row}')
        else:
            ws[f'A{row}'] = param
            ws[f'B{row}'] = value
            ws[f'C{row}'] = notes
        row += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30

def main():
    print("Loading Monte Carlo data...")
    summary, scenarios = load_data()

    print("Creating Excel workbook...")
    wb = Workbook()

    print("  Creating Summary Dashboard...")
    create_summary_sheet(wb, summary)

    print("  Creating Scenario Data sheet...")
    create_scenario_data_sheet(wb, scenarios)

    print("  Creating Distribution Analysis...")
    create_distribution_sheet(wb, scenarios)

    print("  Creating Charts...")
    create_charts_sheet(wb, scenarios)

    print("  Creating Model Assumptions...")
    create_assumptions_sheet(wb)

    print(f"Saving to {OUTPUT_PATH}...")
    wb.save(OUTPUT_PATH)

    print(f"\n✅ Excel workbook created: {OUTPUT_PATH}")
    print("\nSheets included:")
    print("  1. Summary Dashboard - Key metrics with formatting")
    print("  2. Scenario Data - All 10,000 scenarios for analysis")
    print("  3. Distribution Analysis - Percentile buckets and statistics")
    print("  4. Charts - Visual histograms and percentile curves")
    print("  5. Model Assumptions - Complete parameter reference")

if __name__ == "__main__":
    main()
